"""
unificar_ventas.py
------------------
Unifica los reportes anuales de Shares ("Cajas de Sucursales por Período"),
uno por local, en un único master 2025 listo para el comparativo interanual.

Entrada : una carpeta con los .xls de cada local (formato Shares).
          OJO: Shares exporta con extensión .xls pero el archivo es en realidad
          un .xlsx (OOXML/zip). openpyxl lo rechaza por la extensión, así que lo
          abrimos vía BytesIO (que valida el contenido, no el nombre).
Salida  : un .xlsx con dos hojas:
            - "Ventas 2025": una fila por (local, día) con la Venta Bruta.
            - "Resumen"    : total y días por local, con fórmulas SUMAR.SI/CONTAR.SI.

Preferencias del proyecto: stdlib + openpyxl, sin pandas. Fórmulas escritas en
inglés en el XML (openpyxl) pero Excel te las muestra en español.
"""
import argparse
import io
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Mapeo archivo -> (nombre de display, grupo) --------------------------------
# La clave se normaliza (minúsculas, sin tildes, sin separadores) para tolerar
# "Málaga_3", "malaga 3", "Malaga3", etc. Así el mapeo no se rompe por un guion.
# El nombre INTERNO de Shares es más genérico ("MALAGA", "BARCELONA"), por eso
# mapeamos por nombre de ARCHIVO, que respeta tu nomenclatura (Málaga 1 vs 3).
LOCALES = {
    "barcelona1": ("Barcelona 1", "Propia"),
    "barcelona2": ("Barcelona 2", "Propia"),
    "malaga1":    ("Málaga 1",    "Propia"),
    "roma":       ("Roma",        "Propia"),
    "granada":    ("Granada",     "Franquicia"),
    "malaga3":    ("Málaga 3",    "Franquicia"),
    "valencia":   ("Valencia",    "Franquicia"),
    # Madrid y Alicante abrieron en 2026: no tienen histórico 2025.
}

# Orden en que queremos ver los locales en la salida.
ORDEN = ["Roma", "Barcelona 1", "Barcelona 2", "Málaga 1",
         "Granada", "Málaga 3", "Valencia"]

FECHA_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4})")


def clave_local(nombre_archivo):
    """Normaliza el nombre de archivo (sin extensión) a la clave del dict LOCALES."""
    stem = Path(nombre_archivo).stem
    # quito tildes
    stem = "".join(c for c in unicodedata.normalize("NFD", stem)
                   if unicodedata.category(c) != "Mn")
    # dejo solo letras y números en minúscula
    return re.sub(r"[^a-z0-9]", "", stem.lower())


def a_float(v):
    """Convierte a float tolerando formato europeo (1.234,56) si viniera como texto.
    Si openpyxl ya lo entregó como número, lo devuelve tal cual."""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("€", "").replace(" ", "")
    # formato europeo: el punto es separador de miles y la coma es decimal
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def leer_local(path):
    """Devuelve (display, grupo, nombre_interno, filas, total_shares).
    filas = lista de dicts {fecha: date, dia_semana: str, venta: float}
    total_shares = el total que reporta la propia fila TOTALES del archivo (para validar).
    """
    ck = clave_local(path.name)
    if ck not in LOCALES:
        raise SystemExit(
            f"[ERROR] No sé a qué local corresponde '{path.name}'. "
            f"Agregá la clave '{ck}' al dict LOCALES."
        )
    display, grupo = LOCALES[ck]

    # Abrir vía BytesIO para saltear la validación de extensión (.xls que es .xlsx).
    data = io.BytesIO(path.read_bytes())
    wb = load_workbook(data, data_only=True, read_only=True)
    ws = wb.active

    filas = []
    total_shares = None
    header_ok = False
    col_vb = 2  # "Venta Bruta" está en la columna C (índice 2) en el formato Shares

    for r in ws.iter_rows(values_only=True):
        c0 = r[0]
        # Título/subtítulo/headers: los detecto y sigo
        if c0 is None or str(c0).strip() == "":
            continue
        c0s = str(c0).strip()

        if c0s.startswith("Sucursal:"):
            nombre_interno = c0s
            continue
        if c0s == "Fecha":
            header_ok = True
            # confirmo la posición de "Venta Bruta" por las dudas
            fila = list(r)
            if "Venta Bruta" in [str(x) for x in fila]:
                col_vb = [str(x) for x in fila].index("Venta Bruta")
            continue
        if c0s.upper() == "TOTALES":
            total_shares = a_float(r[col_vb])
            continue

        # ¿Es una fila de día? Debe tener una fecha dd/mm/aaaa en la col 0.
        m = FECHA_RE.search(c0s)
        if not m:
            continue  # fila de continuación (2º punto de venta) u otra cosa
        d, mth, y = map(int, m.groups())
        fecha = date(y, mth, d)
        dia_semana = c0s.split()[0]  # "Mié", "Jue", etc.
        filas.append({"fecha": fecha, "dia_semana": dia_semana,
                      "venta": a_float(r[col_vb])})

    if not header_ok:
        raise SystemExit(f"[ERROR] {path.name}: no encontré el header 'Fecha'. "
                         f"¿Cambió el formato de Shares?")
    return display, grupo, filas, total_shares


def construir_master(carpeta, salida):
    archivos = sorted(p for p in Path(carpeta).glob("*.xls"))
    if not archivos:
        raise SystemExit(f"[ERROR] No hay .xls en {carpeta}")

    datos = {}       # display -> {"grupo":..., "filas":[...], "total_shares":...}
    for path in archivos:
        display, grupo, filas, total_shares = leer_local(path)
        # sumo la venta extraída para contrastarla con el total oficial de Shares
        suma = round(sum(f["venta"] for f in filas), 2)
        estado = "OK" if (total_shares is None or abs(suma - total_shares) < 0.5) else "¡DESCUADRE!"
        print(f"  {display:14} [{grupo:10}] {len(filas):3} días | "
              f"extraído {suma:>12,.2f} | Shares {(-1 if total_shares is None else total_shares):>12,.2f} | {estado}")
        datos[display] = {"grupo": grupo, "filas": filas, "total_shares": total_shares}

    # ---- Escribir el master --------------------------------------------------
    wb = Workbook()
    _hoja_datos(wb, datos)
    _hoja_resumen(wb, datos)
    wb.save(salida)
    return datos


# --- Estilos reutilizables -------------------------------------------------------
BORDO = "6B1F2A"        # bordó Lucciano's para headers
DORADO = "C9A227"
GRIS = "F2F2F2"
_thin = Side(style="thin", color="D9D9D9")
BORDE = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _estilo_header(celda):
    celda.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    celda.fill = PatternFill("solid", fgColor=BORDO)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    celda.border = BORDE


def _hoja_datos(wb, datos):
    ws = wb.active
    ws.title = "Ventas 2025"
    encabezados = ["Local", "Grupo", "Fecha", "Día", "Venta Bruta (EUR)"]
    ws.append(encabezados)
    for c in ws[1]:
        _estilo_header(c)

    # una fila por (local, día), respetando ORDEN y luego fecha
    fila_idx = 2
    for display in ORDEN:
        if display not in datos:
            continue
        grupo = datos[display]["grupo"]
        for f in sorted(datos[display]["filas"], key=lambda x: x["fecha"]):
            ws.cell(fila_idx, 1, display).font = Font(name="Arial", size=10)
            ws.cell(fila_idx, 2, grupo).font = Font(name="Arial", size=10)
            cfecha = ws.cell(fila_idx, 3, f["fecha"])
            cfecha.number_format = "dd/mm/yyyy"
            cfecha.font = Font(name="Arial", size=10)
            ws.cell(fila_idx, 4, f["dia_semana"]).font = Font(name="Arial", size=10)
            cv = ws.cell(fila_idx, 5, f["venta"])
            cv.number_format = '#,##0.00" €"'
            cv.font = Font(name="Arial", size=10)
            fila_idx += 1

    ws.freeze_panes = "A2"
    for col, w in zip("ABCDE", (16, 12, 13, 8, 18)):
        ws.column_dimensions[col].width = w
    return ws


def _hoja_resumen(wb, datos):
    ws = wb.create_sheet("Resumen")
    ws.append(["Local", "Grupo", "Días", "Venta Bruta 2025 (EUR)"])
    for c in ws[1]:
        _estilo_header(c)

    n_datos = sum(len(datos[d]["filas"]) for d in datos)  # filas totales en la hoja de datos
    rango_local = f"'Ventas 2025'!$A$2:$A${n_datos + 1}"
    rango_venta = f"'Ventas 2025'!$E$2:$E${n_datos + 1}"

    fila = 2
    presentes = [d for d in ORDEN if d in datos]
    for display in presentes:
        ws.cell(fila, 1, display).font = Font(name="Arial", size=10)
        ws.cell(fila, 2, datos[display]["grupo"]).font = Font(name="Arial", size=10)
        # CONTAR.SI y SUMAR.SI (openpyxl las escribe en inglés; Excel las muestra en español)
        ws.cell(fila, 3, f'=COUNTIF({rango_local},A{fila})').font = Font(name="Arial", size=10)
        cv = ws.cell(fila, 4, f'=SUMIF({rango_local},A{fila},{rango_venta})')
        cv.number_format = '#,##0.00" €"'
        cv.font = Font(name="Arial", size=10)
        fila += 1

    # Total general
    ws.cell(fila, 1, "TOTAL").font = Font(name="Arial", bold=True, size=11)
    ws.cell(fila, 3, f"=SUM(C2:C{fila-1})").font = Font(name="Arial", bold=True, size=11)
    ct = ws.cell(fila, 4, f"=SUM(D2:D{fila-1})")
    ct.number_format = '#,##0.00" €"'
    ct.font = Font(name="Arial", bold=True, size=11)
    ct.fill = PatternFill("solid", fgColor=GRIS)
    ws.cell(fila, 1).fill = PatternFill("solid", fgColor=GRIS)
    ws.cell(fila, 3).fill = PatternFill("solid", fgColor=GRIS)

    for col, w in zip("ABCD", (16, 12, 8, 24)):
        ws.column_dimensions[col].width = w
    return ws


if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Unifica los reportes de Shares en un master 2025.")
    ap.add_argument("carpeta", nargs="?", default=".", help="Carpeta con los .xls (default: actual)")
    ap.add_argument("-o", "--salida", default="Ventas_Europa_2025.xlsx", help="Archivo de salida")
    args = ap.parse_args()

    print(f"Leyendo reportes de: {Path(args.carpeta).resolve()}\n")
    construir_master(args.carpeta, args.salida)
    print(f"\nOK -> {args.salida}")
