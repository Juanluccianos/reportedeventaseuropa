"""
generar_acum2025.py  (Europa)
-----------------------------
Arma el comparativo del año anterior a partir del master 2025 unificado
(data/Ventas_Master_2025.xlsx, hoja "Ventas 2025").

Qué hace:
  1. Lee el día de corte de data/ventas_dia.json (lo produce fetch_shares.py).
  2. Suma del master, por local, la Venta Bruta desde el 1° del mismo mes del año
     anterior hasta ese día (ej. 2025-07-01..2025-07-05).
  3. Escribe data/acum_interanual.json -> {"desde","hasta","acum25":{local: monto}}.

Locales sin datos en ese período (Madrid, Alicante, o Granada antes de abrir)
quedan con acumulado 0.0. El motor (report.py) interpreta el 0 como "sin comparación".
"""
import io
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from locales import BRANCH_ORDER

BASE = Path(__file__).parent
VENTAS_DIA = BASE / "data" / "ventas_dia.json"
MASTER = BASE / "data" / "Ventas_Master_2025.xlsx"
SALIDA = BASE / "data" / "acum_interanual.json"
HOJA = "Ventas 2025"


def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()


def dia_de_corte():
    if not VENTAS_DIA.exists():
        raise SystemExit(f"[ERROR] No existe {VENTAS_DIA}. ¿Corriste fetch_shares.py?")
    data = json.loads(VENTAS_DIA.read_text(encoding="utf-8"))
    return datetime.strptime(data["fecha"], "%Y-%m-%d").date()


def acumular_2025(corte):
    ini = date(corte.year - 1, corte.month, 1)
    fin = date(corte.year - 1, corte.month, corte.day)

    wb = load_workbook(io.BytesIO(MASTER.read_bytes()), data_only=True, read_only=True)
    ws = wb[HOJA]
    header = [str(c) for c in next(ws.iter_rows(values_only=True))]
    col_local = header.index("Local")
    col_fecha = header.index("Fecha")
    col_venta = header.index("Venta Bruta (EUR)")

    acum = {b: 0.0 for b in BRANCH_ORDER}
    for r in ws.iter_rows(min_row=2, values_only=True):
        loc = r[col_local]
        if loc is None:
            continue
        loc = str(loc).strip()
        if loc not in acum:
            continue
        f = _as_date(r[col_fecha])
        if ini <= f <= fin:
            acum[loc] += float(r[col_venta] or 0)

    return ini, fin, {k: round(v, 2) for k, v in acum.items()}


if __name__ == "__main__":
    corte = dia_de_corte()
    ini, fin, acum = acumular_2025(corte)
    SALIDA.write_text(json.dumps(
        {"desde": ini.isoformat(), "hasta": fin.isoformat(), "acum25": acum},
        indent=2, ensure_ascii=False), encoding="utf-8")
    con_datos = [k for k, v in acum.items() if v > 0]
    print(f"OK - acum 2025 {ini}..{fin} | locales con historia: {len(con_datos)}/{len(acum)} "
          f"| total {sum(acum.values()):,.2f} EUR")
