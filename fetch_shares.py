"""
fetch_shares.py  (Europa)   ─── PIEZA PENDIENTE DE AJUSTE CON EL MAIL REAL ───
------------------------------------------------------------------------------
Baja de la casilla (IMAP) los reportes diarios de Shares —un mail por local—,
parsea cada adjunto y escribe el consolidado del día: data/ventas_dia.json
    {"fecha": "2026-07-05", "ventas": {"Roma": 5845.60, "Valencia": 1234.50, ...}}

report.py y generar_acum2025.py leen ESE json, así que quedan totalmente
desacoplados de cómo llegue el mail. Lo único específico del correo vive acá.

QUÉ FALTA CONFIRMAR CON EL PRIMER MAIL REAL (marcado con TODO):
  1. Cómo identificar el local de cada mail: ¿por remitente, por asunto, o por el
     nombre interno del adjunto ("Sucursal: VALENCIA")? -> completar LOCAL_POR_CLAVE.
  2. Si el adjunto diario trae UN día (lo esperado) o un rango; el parser de abajo
     ya toma la fila del día de corte, sirve para ambos casos.
  3. Remitente real de Shares -> SENDER.

El parser del adjunto (parse_adjunto) es el mismo que validamos con el histórico:
Shares exporta .xls que en realidad es .xlsx; "Venta Bruta" está en la columna C;
cada día ocupa 2 filas (2 puntos de venta) y la venta va en la fila con fecha.

Variables de entorno (Secrets):
  IMAP_USER, IMAP_APP_PASS  -> casilla que RECIBE los mails de Shares
"""
import email
import imaplib
import io
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

from locales import BRANCH_ORDER

BASE = Path(__file__).parent
SALIDA = BASE / "data" / "ventas_dia.json"

# TODO(mail real): remitente de Shares
SENDER = os.environ.get("SHARES_SENDER", "no-reply@shares.com")

# TODO(mail real): mapear la "clave" que identifica al local (nombre interno de Shares,
# asunto o remitente) al nombre de display. Las claves se normalizan (sin tildes,
# minúsculas). Ejemplo tentativo según el nombre interno visto en el histórico:
LOCAL_POR_CLAVE = {
    "roma": "Roma",
    "barcelona": "Barcelona 1",
    "barcelona2": "Barcelona 2",
    "madrid": "Madrid",
    "malaga": "Málaga 1",
    "malaga3": "Málaga 3",
    "granada": "Granada",
    "alicante": "Alicante",
    "valencia": "Valencia",
}
FECHA_RE = re.compile(r"(\d{2})/(\d{2})/(\d{4})")


def _norm(s):
    s = "".join(c for c in unicodedata.normalize("NFD", s or "") if unicodedata.category(c) != "Mn")
    return re.sub(r"[^a-z0-9]", "", s.lower())


def a_float(v):
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("€", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_adjunto(payload):
    """Devuelve (clave_local, {fecha_iso: venta}). Toma las filas-día del reporte."""
    wb = load_workbook(io.BytesIO(payload), data_only=True, read_only=True)
    ws = wb.active
    clave, col_vb = None, 2
    ventas = {}
    for r in ws.iter_rows(values_only=True):
        c0 = r[0]
        if c0 is None or str(c0).strip() == "":
            continue
        c0s = str(c0).strip()
        if c0s.startswith("Sucursal:"):
            # "Sucursal: VALENCIA - 01/01/2025 - 31/12/2025" -> clave del nombre
            nombre = c0s.split(":", 1)[1].split(" - ")[0].strip()
            clave = _norm(nombre)
            continue
        if c0s == "Fecha":
            fila = [str(x) for x in r]
            if "Venta Bruta" in fila:
                col_vb = fila.index("Venta Bruta")
            continue
        if c0s.upper() == "TOTALES":
            continue
        m = FECHA_RE.search(c0s)
        if m:
            d, mth, y = map(int, m.groups())
            ventas[date(y, mth, d).isoformat()] = a_float(r[col_vb])
    return clave, ventas


def resolver_local(clave_interna, remitente, asunto):
    """Decide el local. TODO(mail real): ajustar según de dónde salga la identidad."""
    for cand in (clave_interna, _norm(asunto), _norm(remitente)):
        if cand and cand in LOCAL_POR_CLAVE:
            return LOCAL_POR_CLAVE[cand]
    return None


def main():
    user = os.environ["IMAP_USER"]
    pwd = os.environ["IMAP_APP_PASS"]
    # Día objetivo: por defecto, ayer (el cierre que llega a la mañana). Override con FECHA_OBJETIVO.
    objetivo = os.environ.get("FECHA_OBJETIVO")
    objetivo = datetime.strptime(objetivo, "%Y-%m-%d").date() if objetivo else (datetime.utcnow().date() - timedelta(days=1))

    M = imaplib.IMAP4_SSL("imap.gmail.com")
    M.login(user, pwd)
    M.select("INBOX")
    since = (objetivo - timedelta(days=1)).strftime("%d-%b-%Y")
    typ, data = M.search(None, f'(FROM "{SENDER}" SINCE {since})')
    ids = data[0].split()
    if not ids:
        print(f"[ERROR] No hay mails de {SENDER} desde {since}.")
        return 1

    ventas = {}
    for mid in reversed(ids):
        typ, msgdata = M.fetch(mid, "(RFC822)")
        msg = email.message_from_bytes(msgdata[0][1])
        asunto = msg.get("Subject", "")
        remitente = msg.get("From", "")
        for part in msg.walk():
            fn = part.get_filename()
            if not fn or not fn.lower().endswith((".xls", ".xlsx")):
                continue
            clave, por_fecha = parse_adjunto(part.get_payload(decode=True))
            local = resolver_local(clave, remitente, asunto)
            if not local:
                print(f"[AVISO] No pude identificar el local del mail '{asunto}' (clave={clave!r}).")
                continue
            if local in ventas:
                continue  # ya tengo el más reciente de este local
            v = por_fecha.get(objetivo.isoformat())
            if v is None:
                print(f"[AVISO] {local}: el adjunto no trae el día {objetivo}.")
                continue
            ventas[local] = round(v, 2)
    M.logout()

    faltan = [b for b in BRANCH_ORDER if b not in ventas]
    if faltan:
        # Reporte parcial: los que faltan van con lo que haya (0 si nunca llegó) y se avisa.
        print(f"[AVISO] Faltan locales para {objetivo}: {', '.join(faltan)}")

    SALIDA.write_text(json.dumps({"fecha": objetivo.isoformat(), "ventas": ventas},
                                 indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"[OK] {SALIDA.name} -> {len(ventas)}/{len(BRANCH_ORDER)} locales para {objetivo}.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
